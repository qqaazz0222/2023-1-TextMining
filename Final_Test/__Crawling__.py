# [ 라이브러리 ]
import requests as rq
from bs4 import BeautifulSoup
import openpyxl
import time

# [ 엑셀로 저장하기 위한 워크북 생성 ]
wb = openpyxl.Workbook()
# [ 첫번째 시트 선택 ]
ws = wb.active
# [ A1셀에 title입력 ]
ws['A1'] = 'title'

# [ 뉴스 검색 키워드 설정 ]
keyword = "인공지능"
# [ 검색 페이지 수 설정 (1페이지 당 10개의 뉴스) ]
lastPageNo = 400

# [ 초기 페이지 설정 ]
pageNo = 0
# [ 엑셀로 저장할 열 설정 ]
rowNo = 2
# [ 데이터 크롤링 후 엑셀에 저장 ]
for articleNo in range(1, lastPageNo*10, 10):
    pageNo += 1
    print(f'=============={pageNo}페이지==============')
    url = f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&start={articleNo}"
    res = rq.get(url)
    html = res.content.decode("utf-8")
    soup = BeautifulSoup(html, 'lxml')

    words = soup.select('.news_tit')
    for word in words:
        title = word.attrs['title']
        link = word.attrs['href']
        print(title)
        ws[f'A{rowNo}'] = title
        rowNo += 1
# [ 엑셀 파일 저장 ]
wb.save('./ainews.xlsx')