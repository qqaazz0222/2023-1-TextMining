# [ 라이브러리 불러오기 ]
import requests as rq
from bs4 import BeautifulSoup
import openpyxl

# [ 엑셀 파일로 저장 준비 ]
wb = openpyxl.Workbook()
ws = wb.create_sheet('news')
ws['A1'] = "time"
ws['B1'] = "word"

# [ 뉴스 데이터 수집하기 ]
url = [
    "https://search.naver.com/search.naver?where=news&query=인공지능&ds=2020.01.01&de=2022.11.30&nso=so:r,p:from20200101to20221130,a:all&start=",
    "https://search.naver.com/search.naver?where=news&query=인공지능&ds=2022.12.01&de=2023.04.27&nso=so:r,p:from20221201to20230427,a:all&start="
]
lastPageNo = 100
txt = ["", ""]

row_Num = 2
for n in range(2):
    time = "before"
    if n != 0:
        time = "after"
    pageNo = 0
    for articleNo in range(1, lastPageNo*10, 10):
        pageNo += 1
        now_url = url[n] + str(articleNo)
        res = rq.get(now_url)
        html = res.content.decode("utf-8")
        soup = BeautifulSoup(html, 'lxml')
        words = soup.select('.news_tit')
        for word in words:
            pos1 = "A" + str(row_Num) 
            pos2 = "B" + str(row_Num)
            ws[pos1] = time
            ws[pos2] = word.attrs['title']
            txt[n] += word.attrs['title'] + "\n"
            row_Num += 1
        print(f"[Process {n}] {articleNo//10 + 1} / {lastPageNo}, {now_url}")
    print(f"[Process {n}] Done!")

# [ 텍스트 파일로 저장 ]
t_f = open("total.txt", "w")
t_f.write(txt[0] + txt[1])
t_f.close()

# [ 엑셀 파일로 저장 ]
wb.save("news_data.xlsx")
wb = openpyxl.load_workbook("news_data.xlsx")
wb.remove_sheet(wb['Sheet'])
wb.save("news_data.xlsx")